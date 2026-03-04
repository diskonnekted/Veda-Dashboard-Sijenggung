import json
import re
from openpyxl import load_workbook

SRC = r"..\data\penduduk_04_03_2026.xlsx"
OUT = r"codebook.json"

def parse_mapping(cell_text: str):
    text = (cell_text or "").strip()
    if not text:
        return None
    parts = re.split(r"[;,]", text)
    mapping = {}
    for p in parts:
        if ":" in p:
            k, v = p.split(":", 1)
        elif "=" in p:
            k, v = p.split("=", 1)
        else:
            continue
        k = re.sub(r"[^\d\-]+", "", k).strip()
        v = v.strip()
        if k:
            mapping[k] = v
    return mapping or None

def main():
    wb = load_workbook(SRC, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    codes_row = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
    headers_row = [c.value for c in next(ws.iter_rows(min_row=3, max_row=3))]

    codebook = {}
    for idx, header in enumerate(headers_row):
        header = (header or "").strip()
        code_hint = (codes_row[idx] or "") if idx < len(codes_row) else ""
        if not header:
            continue
        m = parse_mapping(str(code_hint))
        if m:
            codebook[header] = m

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(codebook, f, ensure_ascii=False, indent=2)
    print(f"Saved {OUT} with {len(codebook)} mapped columns")

if __name__ == "__main__":
    main()
